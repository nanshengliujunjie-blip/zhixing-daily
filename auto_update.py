#!/usr/bin/env python3
"""
知星全自动更新脚本 - 一键完成所有数据更新
用法: python3 ~/zhixing/auto_update.py

更新内容:
  1. FALLBACK (index.html + dashboard.html) ← update_daily.py 逻辑
  2. user_orders.json ← Nexita dashboard 7 (fetch_orders.mjs)
  3. ad_users.json ← 归因工具 (fetch_attribution.mjs)
  4. consultant_data.json + consultant_all_data.json ← 重建
  5. git commit + push
"""
import json, re, subprocess, sys, os
from datetime import date, timedelta
from pathlib import Path

REPO = Path(__file__).parent
INDEX_HTML     = REPO / 'index.html'
DASHBOARD_HTML = REPO / 'dashboard.html'
AD_USERS       = REPO / 'ad_users.json'
USER_ORDERS    = REPO / 'user_orders.json'
USER_DATA      = REPO / 'user_data.json'
CONSULTANT_D   = REPO / 'consultant_data.json'
CONSULTANT_ALL = REPO / 'consultant_all_data.json'
FETCH_ORDERS   = REPO / 'fetch_orders.mjs'
FETCH_ATTR     = REPO / 'fetch_attribution.mjs'
FETCH_MATERIALS= REPO / 'fetch_materials.mjs'
UPDATE_DAILY   = REPO / 'update_daily.py'
MATERIALS_JSON = REPO / 'materials.json'

today = date.today()
yesterday = today - timedelta(days=1)

print(f'\n{"="*55}')
print(f'  知星全自动更新  {today}')
print(f'{"="*55}\n')

# ═══════════════════════════════════════════════════════
# Step 1: 更新 FALLBACK（复用 update_daily.py）
# 可用 SKIP_FALLBACK=1 跳过（表1表2约中午才更新时，先只更新用户/投放数据）
# ═══════════════════════════════════════════════════════
# 日报FALLBACK改用 compass(表1+表2)，可靠且快：每次都补新日期；
# full模式(非SKIP_FALLBACK)额外回填最近35天历史ROI曲线。
if True:
    print('【Step 1】更新 FALLBACK (compass: 表1+表2)...')
    _env = {**os.environ}
    if os.environ.get('SKIP_FALLBACK'):
        print('  (轻量模式: 只补新日期，跳过历史ROI回填)')
    else:
        _env['FALLBACK_ROI'] = '1'
    result = subprocess.run([sys.executable, str(UPDATE_DAILY)], cwd=REPO, env=_env)
    if result.returncode != 0:
        print('  ⚠ update_daily.py 失败，继续其他步骤')
    else:
        print('  ✓ index.html FALLBACK 已更新')
        # 同步 dashboard.html 的 FALLBACK（与 index.html 保持一致）
        idx = INDEX_HTML.read_text(encoding='utf-8')
        m = re.search(r'(const FALLBACK\s*=\s*\[)(.*?)(\];)', idx, re.DOTALL)
        if m:
            fb_content = m.group(2)
            dash = DASHBOARD_HTML.read_text(encoding='utf-8')
            dash2 = re.sub(r'(const FALLBACK\s*=\s*\[)(.*?)(\];)',
                           lambda x: x.group(1) + fb_content + x.group(3), dash, flags=re.DOTALL)
            DASHBOARD_HTML.write_text(dash2, encoding='utf-8')
            print('  ✓ dashboard.html FALLBACK 已同步')

# ═══════════════════════════════════════════════════════
# Step 2: 增量更新 user_orders.json
# ═══════════════════════════════════════════════════════
print('\n【Step 2】更新 user_orders.json...')

with open(USER_ORDERS) as f:
    orders_data = json.load(f)

# 找最新订单日期
all_dates = set()
for orders in orders_data.values():
    for o in orders:
        all_dates.add(o['d'])
last_order_date = max(all_dates) if all_dates else '2025-04-30'
fetch_to   = today.isoformat()   # 含今天：实时累积今日订单
fetch_from = (date.fromisoformat(last_order_date) + timedelta(days=1)).isoformat()
if fetch_from > fetch_to:
    fetch_from = fetch_to        # 今天已抓过，仍重抓今天以刷新(dedup去重)
if True:
    print(f'  拉取 {fetch_from} ~ {fetch_to} 的订单...')
    proc = subprocess.run(
        ['node', str(FETCH_ORDERS), f'--from={fetch_from}', f'--to={fetch_to}'],
        capture_output=True, text=True, timeout=300, cwd=REPO
    )
    if proc.returncode != 0:
        print(f'  ⚠ fetch_orders.mjs 失败: {proc.stderr[-300:]}')
    else:
        new_orders = 0
        done_info = {}
        for line in proc.stdout.strip().split('\n'):
            if not line: continue
            try:
                obj = json.loads(line)
            except:
                continue
            if obj.get('_done'):
                done_info = obj
                continue
            uid = obj['uid']
            entry = {'d': obj['d'], 'c': obj['c'], 'amt': obj['amt'], 'type': obj['type']}
            if uid not in orders_data:
                orders_data[uid] = []
            # 避免重复（同日期同订单）
            existing = {(o['d'], o.get('amt'), o.get('type')) for o in orders_data[uid]}
            key = (entry['d'], entry['amt'], entry['type'])
            if key not in existing:
                orders_data[uid].append(entry)
                new_orders += 1

        if new_orders > 0:
            with open(USER_ORDERS, 'w') as f:
                json.dump(orders_data, f, ensure_ascii=False, separators=(',', ':'))
            print(f'  ✓ user_orders.json 新增 {new_orders} 条（涉及 {len(orders_data)} 个用户）')
        else:
            print(f'  orders 无新数据')

# ═══════════════════════════════════════════════════════
# Step 3: 增量更新 ad_users.json（归因 Excel）
# ═══════════════════════════════════════════════════════
print('\n【Step 3】更新 ad_users.json...')

import openpyxl, tempfile

with open(AD_USERS) as f:
    ad = json.load(f)
existing_uids = set(str(u[0]) for u in ad['users'])
admap = {str(u[0]): u for u in ad['users']}
existing_reg_dates = set(u[10] for u in ad['users'] if u[10])
last_ad_date = max(existing_reg_dates) if existing_reg_dates else '2025-04-30'

fetch_attr_to   = today.isoformat()   # 含今天：实时累积今日注册/付费
fetch_attr_from = date.fromisoformat(last_ad_date).isoformat()
if fetch_attr_from > fetch_attr_to:
    fetch_attr_from = fetch_attr_to    # 今天已抓过，仍重抓今天以刷新

with open(USER_DATA) as f:
    ud = json.load(f)
ud_map = {str(u[0]): u for u in (ud['users'] if 'users' in ud else ud)}

with open(USER_ORDERS) as f:
    orders_data_fresh = json.load(f)

total_ad_new = 0
paid_fixed = 0
orders_dirty = False
if fetch_attr_from > fetch_attr_to:
    print(f'  ad_users 已是最新 ({last_ad_date})，跳过')
else:
    ds = fetch_attr_from
    while ds <= fetch_attr_to:
        print(f'  下载 {ds} 归因数据...')
        tmp_xlsx = os.path.join(tempfile.gettempdir(), f'attr_{ds}.xlsx')
        r = subprocess.run(
            ['node', str(FETCH_ATTR), f'--date={ds}', f'--out={tmp_xlsx}'],
            capture_output=True, text=True, timeout=180, cwd=REPO
        )
        if r.returncode != 0:
            print(f'    ⚠ 下载失败: {r.stderr[-200:]}')
        else:
            try:
                wb = openpyxl.load_workbook(tmp_xlsx, data_only=True)
                # 优先用「点击归因明细数据」sheet（含首日付费事件）
                ws = wb['点击归因明细数据'] if '点击归因明细数据' in wb.sheetnames else wb.active
                rows_xlsx = list(ws.iter_rows(values_only=True))
                # 解析新增注册 + 首日付费（权威付费源，金额单位 分→元）+ 媒体来源
                new_uids = []
                pay_orders = {}   # uid -> [元金额,...]
                media_map = {}    # uid -> 媒体（优先新增注册事件）
                for r2 in rows_xlsx[1:]:
                    ev = r2[3]; uid = str(r2[16]) if r2[16] else None
                    if not uid or uid == '0': continue
                    if r2[5] and (uid not in media_map or ev == '新增注册'):
                        media_map[uid] = r2[5]
                    if ev == '新增注册':
                        new_uids.append(uid)
                    elif ev == '首日付费' and r2[17]:
                        pay_orders.setdefault(uid, []).append(round(r2[17] / 100.0, 2))

                # 1) 新增注册用户入库（首日付费额取权威首日付费事件，缺则0）
                truly_new = [u for u in new_uids if u not in existing_uids]
                for uid in truly_new:
                    uu = ud_map.get(uid)
                    gender = uu[2] if uu and len(uu) > 2 else 0
                    age    = uu[3] if uu and len(uu) > 3 else None
                    orders = orders_data_fresh.get(uid, [])
                    roi_pay = round(sum(pay_orders.get(uid, [])), 2)
                    paid_cnt = len(pay_orders.get(uid, []))
                    total_amt = round(max(sum(o['amt'] for o in orders), roi_pay), 2)
                    order_cnt = len(orders)
                    q_cnt  = sum(1 for o in orders if o.get('type') == '提问')
                    lm_cnt = sum(1 for o in orders if o.get('type') == '连麦')
                    zx_cnt = sum(1 for o in orders if o.get('type') == '咨询')
                    fd_cons = [o.get('c') for o in orders if o['d'] == ds and o.get('c') and o.get('c') != '未知']
                    consultant = fd_cons[0] if fd_cons else ('知小i' if roi_pay > 0 else '未知')
                    u = [uid, gender, age, roi_pay, total_amt,
                         order_cnt, q_cnt, lm_cnt, zx_cnt, consultant, ds, paid_cnt,
                         media_map.get(uid, '')]
                    ad['users'].append(u)
                    existing_uids.add(uid); admap[uid] = u
                    total_ad_new += 1

                # 2a) 用「真实付费用户统计明细」sheet 补齐 pay_orders（兜底点击归因遗漏的付费）
                if '真实付费用户统计明细' in wb.sheetnames:
                    ws_real = wb['真实付费用户统计明细']
                    rows_real = list(ws_real.iter_rows(values_only=True))
                    if rows_real and len(rows_real) > 1:
                        rh = rows_real[0]
                        r_uid_col = list(rh).index('uid') if 'uid' in rh else None
                        r_amt_col = list(rh).index('付费金额') if '付费金额' in rh else None
                        r_media_col = list(rh).index('媒体') if '媒体' in rh else None
                        if r_uid_col is not None and r_amt_col is not None:
                            real_pay = {}
                            for rr in rows_real[1:]:
                                ruid = str(rr[r_uid_col]) if rr[r_uid_col] else None
                                if not ruid or ruid == '0': continue
                                ramt = round(float(rr[r_amt_col] or 0) / 100.0, 2)
                                if ramt <= 0: continue
                                real_pay.setdefault(ruid, []).append(ramt)
                                if r_media_col is not None and rr[r_media_col] and ruid not in media_map:
                                    media_map[ruid] = rr[r_media_col]
                            for ruid, ramts in real_pay.items():
                                pay_orders[ruid] = ramts

                # 2b) 用首日付费事件校正所有当日付费用户（含已存在用户），避免漏计AI小额付费
                for uid, amts in pay_orders.items():
                    u = admap.get(uid)
                    if not u: continue
                    target = round(sum(amts), 2)
                    cnt = len(amts)
                    if (u[3] or 0) != target:
                        u[3] = target
                        u[4] = round(max(u[4] or 0, target), 2)
                        u[5] = max(u[5] or 0, cnt)
                        if len(u) <= 11: u.append(cnt)
                        else: u[11] = cnt
                        paid_fixed += 1
                    existing = orders_data_fresh.get(uid, [])
                    cur_day = [x for x in existing if x['d'] == ds]
                    cur_paid = round(sum(x['amt'] for x in cur_day if x['amt'] > 0), 2)
                    if abs(cur_paid - target) > 0.01:
                        cc = [x['c'] for x in cur_day if x.get('c') and x['c'] != '未知']
                        consultant = cc[0] if cc else '知小i'
                        g = cur_day[0].get('gender') if cur_day else (u[1] or 0)
                        a = cur_day[0].get('age') if cur_day else u[2]
                        otype = (cur_day[0].get('type') if cur_day else None) or '提问'
                        kept = [x for x in existing if x['d'] != ds]
                        for amt in amts:
                            kept.append({'d': ds, 'c': consultant, 'amt': amt,
                                         'type': otype, 'gender': g, 'age': a})
                        orders_data_fresh[uid] = kept
                        orders_dirty = True

                # 3) 回填媒体来源 u[12]（已存在用户补齐；索引11=付费单数, 12=媒体）
                for uid, media in media_map.items():
                    u = admap.get(uid)
                    if not u or not media: continue
                    if len(u) < 12: u.append(0)      # 补 paid_cnt(11)
                    if len(u) < 13: u.append('')     # 补 media(12)
                    u[12] = media

                print(f'    新增 {len(truly_new)} 人 / 校正付费 {len(pay_orders)} 人')
                os.unlink(tmp_xlsx)
            except Exception as e:
                print(f'    ⚠ 解析失败: {e}')
        ds = (date.fromisoformat(ds) + timedelta(days=1)).isoformat()

if total_ad_new > 0 or paid_fixed > 0:
    ad['users'].sort(key=lambda u: u[10] or '')
    with open(AD_USERS, 'w') as f:
        json.dump(ad, f, ensure_ascii=False, separators=(',', ':'))
    print(f'  ✓ ad_users.json 新增 {total_ad_new} 人, 校正付费 {paid_fixed} 人（共 {len(ad["users"])} 人）')
if orders_dirty:
    with open(USER_ORDERS, 'w') as f:
        json.dump(orders_data_fresh, f, ensure_ascii=False, separators=(',', ':'))
    print(f'  ✓ user_orders.json 已按权威首日付费校正')

# ═══════════════════════════════════════════════════════
# Step 3.5: 回填近期注册用户的 gender/age（来自订单画像，幂等）
# ═══════════════════════════════════════════════════════
print('\n【Step 3.5】回填 gender/age...')
try:
    _genv = {**os.environ}
    for _k in ('HTTP_PROXY','HTTPS_PROXY','http_proxy','https_proxy','ALL_PROXY','all_proxy'):
        _genv.pop(_k, None)
    r = subprocess.run([sys.executable, str(REPO / 'backfill_gender.py')],
                       cwd=REPO, env=_genv, timeout=600)
    if r.returncode != 0:
        print('  ⚠ backfill_gender 失败，继续')
except Exception as e:
    print(f'  ⚠ backfill_gender 异常: {e}，继续')

# ═══════════════════════════════════════════════════════
# Step 3.6: 捡漏付费（订单系统首日付费 > 记录值则补，针对xlsx漏记大单）
# ═══════════════════════════════════════════════════════
print('\n【Step 3.6】捡漏付费...')
try:
    r = subprocess.run([sys.executable, str(REPO / 'reconcile_payments.py')],
                       cwd=REPO, timeout=120)
    if r.returncode != 0:
        print('  ⚠ reconcile_payments 失败，继续')
except Exception as e:
    print(f'  ⚠ reconcile_payments 异常: {e}，继续')

# ═══════════════════════════════════════════════════════
# Step 4: 重建 consultant_data.json + consultant_all_data.json
# ═══════════════════════════════════════════════════════
print('\n【Step 4】重建 consultant 数据...')

with open(USER_ORDERS) as f:
    orders_data_final = json.load(f)
with open(AD_USERS) as f:
    ad_final = json.load(f)

# 投放用户 uid → reg_date
ad_uid_to_regdate = {str(u[0]): u[10] for u in ad_final['users'] if u[10]}

# consultant_data: 仅首日投放用户
daily = {}   # {date: {consultant: {amt, users: set, orders}}}
for uid, orders in orders_data_final.items():
    reg_date = ad_uid_to_regdate.get(uid)
    if not reg_date: continue
    for o in orders:
        if o['d'] != reg_date: continue   # 仅首日
        if o.get('amt', 0) <= 0: continue # 付费金额=0不计入付费统计
        cons = o.get('c', '未知') or '未知'
        d = o['d']
        if d not in daily: daily[d] = {}
        if cons not in daily[d]: daily[d][cons] = {'amt': 0, 'users': set(), 'orders': 0}
        daily[d][cons]['amt'] += o['amt']
        daily[d][cons]['users'].add(uid)
        daily[d][cons]['orders'] += 1

daily_out = {}
for d, cons_map in daily.items():
    daily_out[d] = sorted(
        [{'name': c, 'amt': round(v['amt'], 2), 'users': len(v['users']), 'orders': v['orders']}
         for c, v in cons_map.items()],
        key=lambda x: -x['amt']
    )

# totals
totals = {}
for d, rows in daily_out.items():
    for r in rows:
        c = r['name']
        if c not in totals: totals[c] = {'amt': 0, 'users': 0, 'orders': 0}
        totals[c]['amt'] = round(totals[c]['amt'] + r['amt'], 2)
        totals[c]['users'] += r['users']
        totals[c]['orders'] += r['orders']

with open(CONSULTANT_D, 'w') as f:
    json.dump({'daily': daily_out, 'totals': totals}, f, ensure_ascii=False, separators=(',', ':'))
print(f'  ✓ consultant_data.json: {len(daily_out)} 天, {len(totals)} 个顾问')

# consultant_all_data: 所有付费用户（不限投放）
daily_all = {}
for uid, orders in orders_data_final.items():
    for o in orders:
        if o['amt'] <= 0: continue
        cons = o.get('c', '未知') or '未知'
        d = o['d']
        if d not in daily_all: daily_all[d] = {}
        if cons not in daily_all[d]: daily_all[d][cons] = {'amt': 0, 'users': set(), 'orders': 0}
        daily_all[d][cons]['amt'] += o['amt']
        daily_all[d][cons]['users'].add(uid)
        daily_all[d][cons]['orders'] += 1

daily_all_out = {}
for d, cons_map in daily_all.items():
    daily_all_out[d] = sorted(
        [{'name': c, 'amt': round(v['amt'], 2), 'users': len(v['users']), 'orders': v['orders']}
         for c, v in cons_map.items()],
        key=lambda x: -x['amt']
    )

totals_all = {}
for d, rows in daily_all_out.items():
    for r in rows:
        c = r['name']
        if c not in totals_all: totals_all[c] = {'amt': 0, 'users': 0, 'orders': 0}
        totals_all[c]['amt'] = round(totals_all[c]['amt'] + r['amt'], 2)
        totals_all[c]['users'] += r['users']
        totals_all[c]['orders'] += r['orders']

with open(CONSULTANT_ALL, 'w') as f:
    json.dump({'daily': daily_all_out, 'totals': totals_all}, f, ensure_ascii=False, separators=(',', ':'))
print(f'  ✓ consultant_all_data.json: {len(daily_all_out)} 天, {len(totals_all)} 个顾问')

# ═══════════════════════════════════════════════════════
# Step 4.5: 更新素材数据（最近 30 天）
# ═══════════════════════════════════════════════════════
print('\n【Step 4.5】更新推广素材数据...')
try:
    mat_proc = subprocess.run(
        ['node', str(FETCH_MATERIALS), '--days=30'],
        cwd=REPO, capture_output=True, timeout=300
    )
    if mat_proc.returncode == 0 and mat_proc.stdout:
        import json as _json
        _mat = _json.loads(mat_proc.stdout)
        _items = _mat.get('list', [])
        if _items:
            with open(MATERIALS_JSON, 'wb') as f:
                f.write(mat_proc.stdout)
            print(f'  ✓ materials.json: {len(_items)} 条，{_mat.get("startDate")}~{_mat.get("endDate")}')
        else:
            print('  ⚠ 本次素材结果为空，保留原 materials.json')
    else:
        print(f'  ⚠ fetch_materials 失败: {mat_proc.stderr.decode()[-300:]}')
except Exception as e:
    print(f'  ⚠ 素材更新异常: {e}，继续')

# ═══════════════════════════════════════════════════════
# Step 5: Git commit + push
# ═══════════════════════════════════════════════════════
print('\n【Step 5】Git commit & push...')
files = ['index.html', 'dashboard.html', 'user_orders.json', 'ad_users.json',
         'consultant_data.json', 'consultant_all_data.json', 'materials.json']
subprocess.run(['git', 'add'] + files, cwd=REPO, check=True)
msg = f'全量自动更新 {today}: orders+ad_users+consultant+FALLBACK'

# 先提交（无变更则跳过）
committed = subprocess.run(['git', 'commit', '-m', msg], cwd=REPO).returncode == 0
if not committed:
    print('  (无变更，跳过提交)')

# 再推送，单独捕获 push 失败并给出可操作提示
push = subprocess.run(
    ['git', 'push'], cwd=REPO,
    env={**os.environ, 'GIT_TERMINAL_PROMPT': '0'},
    capture_output=True, text=True,
)
if push.returncode == 0:
    print('✅ 已推送到 GitHub Pages!')
else:
    err = (push.stderr or '') + (push.stdout or '')
    if any(k in err for k in ('could not read Username', 'Authentication failed',
                               'terminal prompts disabled', '403', 'invalid credentials')):
        print('❌ Push 失败：缺少 GitHub 凭据。')
        print('   本地已移除明文 token，需录入新 token：')
        print('   在仓库目录手动执行一次  git push  ，按提示输入')
        print('   用户名 = GitHub 账号，密码 = 新生成的 Personal Access Token(勾 repo 权限)')
        print('   录入一次后 macOS 钥匙串会记住，后续自动推送即恢复正常。')
        print('   ⚠️ 本次数据已在本地更新并提交，待凭据修好后再次运行即会推送。')
    else:
        print('❌ Push 失败：')
        print('   ' + err.strip().replace('\n', '\n   '))

print('\n✅ 全部更新完成！')
