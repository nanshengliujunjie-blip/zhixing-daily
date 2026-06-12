#!/usr/bin/env python3
"""一次性回填 ad_users 媒体来源 u[12]，源自已有归因 Excel（5/10~6/09）。"""
import openpyxl, json
from collections import defaultdict, Counter
from pathlib import Path

REPO = Path(__file__).parent
FILES = [
    '/Users/oulei/Downloads/MTc4MTA4Mjc5NDQ2NCM1MzkjeGxzeA==.xlsx',  # 5/10~6/09 月度
    '/Users/oulei/Downloads/MTc4MTA1NjM5ODEwNyMyNTkjeGxzeA==.xlsx',  # 6/09 单日
]
uid_media = defaultdict(Counter)
for f in FILES:
    if not Path(f).exists(): continue
    wb = openpyxl.load_workbook(f, read_only=True)
    ws = wb['点击归因明细数据'] if '点击归因明细数据' in wb.sheetnames else wb.active
    for r in ws.iter_rows(min_row=2, values_only=True):
        uid = str(r[16]) if r[16] else None
        if not uid or uid == '0' or not r[5]: continue
        # 新增注册的媒体权重更高
        uid_media[uid][r[5]] += (10 if r[3] == '新增注册' else 1)
media = {u: c.most_common(1)[0][0] for u, c in uid_media.items()}
print(f'媒体映射: {len(media)} 个 uid')

ad = json.load(open(REPO / 'ad_users.json'))
filled = 0
for u in ad['users']:
    m = media.get(str(u[0]))
    if not m: continue
    if len(u) < 12: u.append(0)
    if len(u) < 13: u.append('')
    if u[12] != m:
        u[12] = m; filled += 1
json.dump(ad, open(REPO / 'ad_users.json', 'w'), ensure_ascii=False, separators=(',', ':'))
print(f'已回填媒体: {filled} 个用户')
# 统计覆盖
have = sum(1 for u in ad['users'] if len(u) > 12 and u[12])
print(f'ad_users 媒体覆盖: {have}/{len(ad["users"])}')
